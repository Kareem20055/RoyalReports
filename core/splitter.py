from openpyxl import load_workbook, Workbook
from models.employee import Employee
from models.report import Report
import os


class ReportSplitter:

    def split(self, file_path):

        wb = load_workbook(file_path)
        ws = wb.active

        header_row = None
        name_col = None

        for r in range(1, 15):

            for c in range(1, ws.max_column + 1):

                value = ws.cell(r, c).value

                if value is None:
                    continue

                text = str(value).strip()

                if text == "الاسم" or text.lower() == "name":

                    header_row = r
                    name_col = c
                    break

            if header_row:
                break

        if header_row is None:
            raise Exception("لم يتم العثور على عمود الاسم")

        data_row = header_row + 1

        output = Workbook()
        output.remove(output.active)

        employees = {}
        report = Report()

        for row in ws.iter_rows(min_row=data_row):

            employee = row[name_col - 1].value

            if employee is None:
                continue

            employee = str(employee).strip()

            if employee == "":
                continue

            if employee not in employees:

                sheet = output.create_sheet(employee[:31])

                for r in range(1, header_row + 1):
                    for c in range(1, ws.max_column + 1):
                        sheet.cell(r, c).value = ws.cell(r, c).value

                emp = Employee(name=employee)

                employees[employee] = {
                    "sheet": sheet,
                    "row": header_row + 1,
                    "employee": emp
                }

                report.employees.append(emp)

            sheet = employees[employee]["sheet"]
            current = employees[employee]["row"]

            for c in range(1, ws.max_column + 1):
                sheet.cell(current, c).value = ws.cell(row[0].row, c).value

            row_values = [
                ws.cell(row[0].row, c).value
                for c in range(1, ws.max_column + 1)
            ]

            employees[employee]["employee"].rows.append(row_values)

            employees[employee]["row"] += 1

        output_path = os.path.join(
            os.path.dirname(file_path),
            "Weekly_Separated_By_Employee.xlsx"
        )

        output.save(output_path)

        report.excel_path = output_path

        return report