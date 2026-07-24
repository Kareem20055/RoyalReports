from openpyxl import load_workbook, Workbook
from models.employee import Employee
from models.report import Report
from models.attendance_row import AttendanceRow

import os


class ReportSplitter:

    def split(self, file_path):

        wb = load_workbook(file_path)
        ws = wb.active

        header_row = None

        # البحث عن صف العناوين
        for r in range(1, 15):

            values = []

            for c in range(1, ws.max_column + 1):

                value = ws.cell(r, c).value

                if value is None:
                    values.append("")
                else:
                    values.append(str(value).strip())

            if "الاسم" in values or "Name" in values:
                header_row = r
                break

        if header_row is None:
            raise Exception("لم يتم العثور على صف العناوين")

        # إنشاء قاموس أسماء الأعمدة
        headers = {}

        for c in range(1, ws.max_column + 1):

            value = ws.cell(header_row, c).value

            if value is None:
                continue

            headers[str(value).strip()] = c

        data_row = header_row + 1

        output = Workbook()
        output.remove(output.active)

        employees = {}
        report = Report()

        for row in ws.iter_rows(min_row=data_row):

            name = row[headers["الاسم"] - 1].value

            if name is None:
                continue

            name = str(name).strip()

            if not name:
                continue

            if name not in employees:

                sheet = output.create_sheet(name[:31])

                # نسخ رأس التقرير
                for r in range(1, header_row + 1):
                    for c in range(1, ws.max_column + 1):
                        sheet.cell(r, c).value = ws.cell(r, c).value

                employee = Employee(
                    name=name,
                    person_id=str(row[headers["معرف الشخص"] - 1].value or ""),
                    department=str(row[headers["القسم"] - 1].value or ""),
                    position=str(row[headers["المنصب"] - 1].value or "")
                )

                employees[name] = {
                    "sheet": sheet,
                    "row": header_row + 1,
                    "employee": employee
                }

                report.employees.append(employee)

            sheet = employees[name]["sheet"]
            current_row = employees[name]["row"]

            # نسخ الصف إلى ملف الموظف
            for c in range(1, ws.max_column + 1):
                sheet.cell(current_row, c).value = ws.cell(row[0].row, c).value

            attendance = AttendanceRow(
                date=str(row[headers["التاريخ"] - 1].value or ""),
                weekday=str(row[headers["يوم من الأسبوع"] - 1].value or ""),
                schedule=str(row[headers["جدول المواعيد"] - 1].value or ""),
                check_in=str(row[headers["أول حضور"] - 1].value or ""),
                check_out=str(row[headers["آخر انصراف"] - 1].value or "")
            )

            employees[name]["employee"].rows.append(attendance)
            employees[name]["row"] += 1

        output_path = os.path.join(
            os.path.dirname(file_path),
            "Weekly_Separated_By_Employee.xlsx"
        )

        output.save(output_path)

        report.excel_path = output_path

        return report